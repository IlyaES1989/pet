""" Report classes"""
from io import BytesIO

from openpyxl import (
    load_workbook,
    Workbook,
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from ..report_scripts.weekly.chart import create_chart_sh


class Report:

    def __init__(self):
        self.path = ''
        self.file = ''
        self.workbook = ''
        self.active_sheet = ''

    def create_book(self):
        self.workbook = Workbook()

    def create_sheet(self, titles_sh):
        self.workbook.create_sheet(titles_sh)
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def open_book(self, file_name=None):
        if file_name is None:
            file_name = '%s\\%s' % (self.path, self.file)
        self.workbook = load_workbook(filename=file_name)

    def activate_sheet(self, title_sh):
        self.active_sheet = self.workbook[title_sh]

    def change_height(self, spec_height=None, custom_height=15.75, max_row=200):
        rows = self.active_sheet.row_dimensions
        if spec_height:
            for i in range(max_row):
                if str(i) not in spec_height.keys():
                    rows[i].height = custom_height
                else:
                    rows[i].height = spec_height[str(i)]

        else:
            for i in range(max_row):
                rows[i].height = custom_height

    def change_width(self, spec_width=None, custom_width=11, max_column=100):
        cor_factor = 0.71
        columns = self.active_sheet.column_dimensions

        if spec_width:
            for i in range(1, max_column):
                column_id = get_column_letter(i)
                if column_id not in spec_width.keys():
                    columns[column_id].width = custom_width + cor_factor
                else:
                    columns[column_id].width = spec_width[column_id] + cor_factor
        else:
            for i in range(1, max_column):
                column_id = get_column_letter(i)
                columns[column_id].width = custom_width + cor_factor

    def group_rows(self, rows_dict):
        if rows_dict:
            for key in rows_dict.keys():
                self.active_sheet.row_dimensions.group(int(key), rows_dict[key], hidden=True)

    def group_columns(self, columns_dict):
        if columns_dict:
            for key in columns_dict.keys():
                self.active_sheet.column_dimensions.group(int(key), columns_dict[key], hidden=True)

    def create_merge(self, merge_dict):
        if merge_dict:
            for key in merge_dict.keys():
                for merged in self.active_sheet.merged_cells.ranges:
                    if key in merged:
                        self.active_sheet.merged_cells.remove('%s' % merged)
                self.active_sheet.merge_cells('%s:%s' % (key, merge_dict[key]))

    def write_values(self, cells_dict):
        if cells_dict:
            for key in cells_dict:
                self.active_sheet[key] = cells_dict[key][0]

                if cells_dict[key][1].name in self.workbook.named_styles:
                    self.active_sheet[key].style = cells_dict[key][1].name
                else:
                    self.active_sheet[key].style = cells_dict[key][1]

    def insert_image(self, image_dict):
        if image_dict:
            for key in image_dict:
                img = Image(image_dict[key]['f_name'])
                img.width = image_dict[key]['w']
                img.height = image_dict[key]['h']
                self.active_sheet.add_image(img, key)

    def save(self, name):
        self.workbook.save(name)


class GenerateReport:
    def __init__(self, sheets):
        self.sheets = sheets

    def generate_template(self, year):
        virtual_workbook = BytesIO()
        sheet_num = 0
        report = Report()
        for sheet in self.sheets:
            sheet_num += 1
            sheet.create_template()
            if sheet_num == 1:
                report.create_book()

            report.create_sheet(sheet.sheet_name)
            report.activate_sheet(sheet.sheet_name)
            report.change_height(sheet.high)
            report.change_width(sheet.width)
            report.group_rows(sheet.rows_group)
            report.group_columns(sheet.columns_group)
            report.write_values(sheet.cells)
            report.create_merge(sheet.merge)
            report.insert_image(sheet.image)

            if sheet.chart:
                create_chart_sh(report.workbook['%s' % sheet.sheet_name], year, sheet.chart_title)

            if sheet_num == len(self.sheets):
                report.save(virtual_workbook)

        return virtual_workbook, self.sheets[0].new_report

    def generate_mutable_part(self, last_report):
        virtual_workbook = BytesIO()
        sheet_num = 0
        report = Report()
        for sheet in self.sheets:
            sheet_num += 1
            sheet.create_mutable_part()

            if sheet_num == 1:
                report.open_book(last_report)

            report.activate_sheet(sheet.sheet_name)
            report.write_values(sheet.cells)
            report.create_merge(sheet.merge)
            if sheet_num == len(self.sheets):
                report.save(virtual_workbook)

        return virtual_workbook, self.sheets[0].new_report
