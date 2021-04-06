"""Data class"""
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter
)


def get_max_row(coord_list):
    """ Return max number of rows from list. """
    row = set()
    for element in coord_list:
        number = ''
        list(element)
        for i in element:
            try:
                int(i)
                number += ''.join(str(i))
            except ValueError:
                pass
        row.add(int(number))
    return max(row)


class Data:
    def __init__(self):
        self.high = {}
        self.width = {}
        self.image = {}
        self.rows_group = {}
        self.columns_group = {}
        self.merge = {}
        self.cells = {}

    def add_cells_by_rows(self, data, styles, start_row=1, start_column=1, step=1):
        """
        Write value of date row-by-row to self.cells.
        :param data: list, include other lists with value to write.
        :param styles: list, include other lists with NamedStyle objects.
        :param start_row: int, number of row to start writing.
        :param start_column: int or str, number or letter of column to start writing.
        :param step: int, add gap between rows.
        :return: None
        """
        if len(data) != len(styles):
            raise IndexError('The lengths of data and styles must be equal.')
        else:
            cache = []
            value = data
            styles = styles[:]
            max_len = max([len(value[i]) for i in range(len(value))])

            if type(start_column) == int:
                start_column = get_column_letter(start_column)

            for i in range(len(value)):
                if len(value[i]) != max_len:
                    value[i] = value[i]*(int(max_len/len(value[i])))

            for i in range(len(value[0])):
                for j in range(len(value)):
                    cache.append([value[j][i], styles[j]])

            for i in range(len(cache)):
                if i != 0:
                    start_row = start_row + step
                cell = '%s%s' % (start_column, start_row)
                self.cells[cell] = cache[i]

    def add_cells_by_column(self, data, styles, start_row=1, start_column=1, count_iter=1, step=1):
        """
        Write value of date column-by-column to self.cells.
        :param data: list, include other lists with value to write.
        :param styles: list, include other lists with NamedStyle objects.
        :param start_row: int, number of row to start writing.
        :param start_column: int or str, number or letter of column to start writing.
        :param step: int, add gap between rows.
        :return: None
        """
        if len(data) != len(styles):
            raise IndexError('The lengths of data and styles must be equal.')
        else:
            value = data[:]
            styles = styles[:]
            if type(start_column) == str:
                start_column = column_index_from_string(start_column)

            for i in range(len(value)):
                column = start_column + i*step
                for j in range(count_iter):
                    column = column + j*(len(value))
                    column_letter = get_column_letter(column)
                    cell = '%s%s' % (column_letter, start_row)
                    self.cells[cell] = [value[i], styles[i]]

    def add_merge(self, start_merge, end_merge, count_iter=1, step=1, direction='rightward'):
        """
        Write the range of merge cells to self.merge
        :param start_merge: str, coordinate of the start in the first iteration.
        :param end_merge:str, coordinate of the end in the first iteration.
        :param count_iter: int, the amount of iterations.
        :param step: int, the amount empty rows or columns between each iteration.
        :param direction:str, must be "rightward" or "downward".
        :return:None
        """
        start_column_letter = coordinate_from_string(start_merge)[0]
        start_row = coordinate_from_string(start_merge)[1]
        end_column_letter = coordinate_from_string(end_merge)[0]
        end_row = coordinate_from_string(end_merge)[1]

        if direction == 'rightward':
            start_column = column_index_from_string(start_column_letter)
            end_column = column_index_from_string(end_column_letter)
            len_merge = end_column - start_column

            for i in range(count_iter):
                if i != 0:
                    start_column = start_column + len_merge + step

                end_column = start_column + len_merge
                start_column_letter = get_column_letter(start_column)
                end_column_letter = get_column_letter(end_column)
                self.merge['%s%s' % (start_column_letter, start_row)] = '%s%s' % (end_column_letter, end_row)

        elif direction == 'downward':
            len_merge = end_row - start_row

            for i in range(count_iter):
                if i != 0:
                    start_row = start_row + len_merge + step
                end_row = start_row + len_merge
                self.merge['%s%s' % (start_column_letter, start_row)] = '%s%s' % (end_column_letter, end_row)
        else:
            raise AttributeError('The direction attribute must be "rightward" or "downward"')

    def add_width(self, width_list, start_column=1, count_iter=1):
        """ Write the value of width to self.width"""

        column = start_column
        for i in range(count_iter):
            for width in width_list:
                column_letter = get_column_letter(column)
                self.width['%s' % column_letter] = width
                column += 1



