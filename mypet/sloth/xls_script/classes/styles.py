from openpyxl.styles import (
    NamedStyle,
    Font,
    Border,
    Side,
    PatternFill,
    Alignment,
)

from copy import copy


st = {}

test = NamedStyle(name='test_style')

test.font = Font(name='Times New Roman', bold=True, size=10, color='132763')
test.fill = PatternFill(fill_type='solid', start_color='4A7788')
bd = Side(style='thick', color="DFCECE")
test.border = Border(left=bd, top=bd, right=bd, bottom=bd)
test.alignment = Alignment(horizontal='general', vertical='bottom')
st['test'] = test

base = NamedStyle(name='base_style')
base.font = Font(name='Calibri Light', bold=False, size=11, color='132763')
base.fill = PatternFill(fill_type='solid', start_color='FFFFFF')
bd = Side(style='thick', color="FFFFFF")
base.border = Border(left=bd, top=bd, right=bd, bottom=bd)
base.alignment = Alignment(vertical='bottom')
st['base'] = base

blue20 = copy(base)
blue20.name = 'blue20'
blue20.font.bold = True
blue20.fill.start_color = 'C6D9E0'
blue20.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['blue20'] = blue20

blue20_r = copy(blue20)
blue20_r.name = 'blue20_r'
blue20_r.alignment = Alignment(horizontal='right', vertical='bottom', wrapText=True)
st['blue20_r'] = blue20_r

blue20_p = copy(blue20_r)
blue20_p.name = 'blue20_p'
blue20_p.font.italic = True
blue20_p.number_format = '0.0%'
st['blue20_p'] = blue20_p

blue20_red_p = copy(blue20_p)
blue20_red_p.name = 'blue20_red_p'
blue20_red_p.font.color = 'FF0000'
blue20_red_p.number_format = '0.0%'
st['blue20_red_p'] = blue20_red_p

purple60 = copy(blue20)
purple60.name = 'purple60'
purple60.fill.start_color = 'CBB1B0'
st['purple60'] = purple60

blue05 = copy(blue20_r)
blue05.name = 'blue05'
blue05.fill.start_color = 'DBE9F2'
blue05.font.size = 11
blue05.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['blue05'] = blue05

blue05_r = copy(blue05)
blue05_r.name = 'blue05_r'
blue05_r.alignment = Alignment(horizontal='right')
st['blue05_r'] = blue05_r

blue05_l = copy(blue05)
blue05_l.name = 'blue05_l'
blue05_l.alignment = Alignment(horizontal='left')
st['blue05_l'] = blue05_l

blue10 = copy(blue05)
blue10.name = 'blue10'
blue10.fill.start_color = 'B8D2E7'
blue10.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['blue10'] = blue10

blue10_r = copy(blue10)
blue10_r.name = 'blue10_r'
blue10_r.alignment = Alignment(horizontal='right')
st['blue10_r'] = blue10_r

blue10_l = copy(blue10)
blue10_l.name = 'blue10_l'
blue10_l.alignment = Alignment(horizontal='left')
st['blue10_l'] = blue10_l

blue40_white = copy(blue20)
blue40_white.name = 'blue40_white'
blue40_white.fill.start_color = '4F91C3'
blue40_white.font.size = 14
blue40_white.font.color = 'FFFFFF'
st['blue40_white'] = blue40_white

blue40_white_11 = copy(blue40_white)
blue40_white_11.name = 'blue40_white_11'
blue40_white_11.font.size = 11
blue40_white_11.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['blue40_white_11'] = blue40_white_11

blue40_white_l = copy(blue40_white_11)
blue40_white_l.name = 'blue40_white_l'
blue40_white_l.alignment = Alignment(horizontal='left')
st['blue40_white_l'] = blue40_white_l

blue40_white_11_p = copy(blue40_white)
blue40_white_11_p.name = 'blue40_white_11_p'
blue40_white_11_p.font.size = 11
blue40_white_11_p.font.italic = True
blue40_white_11_p.number_format = '0.0%'
blue40_white_11_p.alignment = Alignment(horizontal='right')
st['blue40_white_11_p'] = blue40_white_11_p

white_red = copy(base)
white_red.name = 'white_red'
white_red.font.bold = True
white_red.font.color = 'C00000'
white_red.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['white_red'] = white_red

white_blue = copy(white_red)
white_blue.name = 'white_blue'
white_blue.font.color = '132763'
white_blue.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['white_blue'] = white_blue

white_blue_l = copy(white_blue)
white_blue_l.name = 'white_blue_l'
white_blue_l.alignment = Alignment(horizontal='left')
st['white_blue_l'] = white_blue_l

white_right = copy(base)
white_right.name = 'white_right'
white_right.alignment = Alignment(horizontal='right')
st['white_right'] = white_right

white_right_p = copy(white_right)
white_right.name = 'white_right_p'
white_right_p.number_format = '0.0%'
white_right_p.font.italic = True
st['white_right_p'] = white_right_p

white_right_red_p = copy(white_right_p)
white_right_red_p.name = 'white_right_red_p'
white_right_red_p.number_format = '0.0%'
white_right_red_p.font.color = 'FF0000'
st['white_right_red_p'] = white_right_red_p

smoky20 = copy(base)
smoky20.name = 'smoky20'
smoky20.fill.start_color = 'F4EEEE'
smoky20.font.bold = True
st['smoky20'] = smoky20

smoky20_v = copy(smoky20)
smoky20_v.name = 'smoky20_v'
smoky20_v.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
st['smoky20_v'] = smoky20_v

smoky20_n = copy(smoky20)
smoky20_n.name = 'smoky20_n'
smoky20_n.number_format = '0'
st['smoky20_n'] = smoky20_n

smoky20_p = copy(smoky20)
smoky20_p.name = 'smoky20_p'
smoky20_p.number_format = '0.0%'
smoky20_p.font.italic = True
st['smoky20_p'] = smoky20_p

smoky20_red_p = copy(smoky20_p)
smoky20_red_p.name = 'smoky20_red_p'
smoky20_red_p.number_format = '0.0%'
smoky20_red_p.font.color = 'FF0000'
st['smoky20_red_p'] = smoky20_red_p

apricot20 = copy(blue20)
apricot20.name = 'apricot20'
apricot20.font.size = 11
apricot20.fill.start_color = 'FEF4E9'
st['apricot20'] = apricot20

apricot20_r = copy(apricot20)
apricot20_r.name = 'apricot20_r'
apricot20_r.alignment = Alignment(horizontal='right', vertical='center', wrapText=True)
st['apricot20_r'] = apricot20_r

apricot20_l = copy(apricot20)
apricot20_l.name = 'apricot20_l'
apricot20_l.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
st['apricot20_l'] = apricot20_l

apricot40 = copy(apricot20)
apricot40.name = 'apricot40'
apricot40.fill.start_color = 'FEEAD1'
st['apricot40'] = apricot40

apricot40_r = copy(apricot40)
apricot40_r.name = 'apricot40_r'
apricot40_r.alignment = Alignment(horizontal='right', vertical='center', wrapText=True)
st['apricot40_r'] = apricot40_r

apricot40_l = copy(apricot40)
apricot40_l.name = 'apricot40_l'
apricot40_l.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
st['apricot40_l'] = apricot40_l

apricot60 = copy(apricot20)
apricot60.name = 'apricot60'
apricot60.fill.start_color = 'FDDFBB'
st['apricot60'] = apricot60

apricot60_r = copy(apricot60)
apricot60_r.name = 'apricot60_r'
apricot60_r.alignment = Alignment(horizontal='right', vertical='center', wrapText=True)
st['apricot60_r'] = apricot60_r

apricot60_l = copy(apricot60)
apricot60_l.name = 'apricot60_l'
apricot60_l.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
st['apricot60_l'] = apricot60_l

apricot60_p = copy(apricot60)
apricot60_p.name = 'apricot60_p'
apricot60_p.font.italic = True
apricot60_p.number_format = '0.0%'
apricot60_p.alignment = Alignment(horizontal='general', vertical='center', wrapText=True)
st['apricot60_p'] = apricot60_p

apricot70 = copy(apricot20)
apricot70.name = 'apricot70'
apricot70.fill.start_color = 'EDB378'
st['apricot70'] = apricot70

apricot80 = copy(apricot20)
apricot80.name = 'apricot80'
apricot80.fill.start_color = 'EDB678'
st['apricot80'] = apricot80

good_p = copy(apricot60_p)
good_p.name = 'good_p'
good_p.fill.start_color = 'C6EFCE'
good_p.font.color = '006100'
good_p.font.italic = True
good_p.number_format = '0.0%'
st['good_p'] = good_p

bad_p = copy(apricot60_p)
bad_p.name = 'bad_p'
bad_p.fill.start_color = 'FFC7CE'
bad_p.font.color = '9C0006'
bad_p.font.italic = True
bad_p.number_format = '0.0%'
st['bad_p'] = bad_p
