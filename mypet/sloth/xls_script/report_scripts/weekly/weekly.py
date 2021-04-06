"""
Weekly report classes.
"""

import os
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl.utils.cell import coordinate_from_string

from ...classes.styles import st
from ...classes.prepare import Prepare
from ...classes.data import (
    Data,
    get_max_row,
)

from ..report_setting import (
    CLUSTER_GROUP,
    PLAN_GROUP,
    TOWN_GROUP,
    CITIES_LIST,
    FULL_CITIES_LIST,
    SUB_CLUSTER_LIST,
)

SCRIPT_DIR = Path(__file__).resolve().parent
IMG_DIR = os.path.join(SCRIPT_DIR, 'image')


def percent(numerator,
            denominator,
            city_dict,
            numerator_correction=1.0,
            denominator_correction=1.0):
    """
    :param numerator: dict, the keys must correspond to city_dict's objects,
     and the values must be numeric.
    :param denominator:dict, the keys must correspond to city_dict's objects,
     and the values must be numeric.
    :param city_dict: any iterable object (dict, set, list and etc.), each iterate of which
    must return the key of numerator and denominator.
    :param numerator_correction: numeric, allowed to increase or decrease each object of numerator.
    :param denominator_correction:numeric, allowed to increase or decrease each object of denominator.
    :return: dictionary, where the keys correspond to city_dict objects
            the values include ratio the numerators to denominators (type: float).
    """
    percent_dict = {}
    for key in city_dict:
        n = numerator.get(key, 0) * numerator_correction
        d = denominator.get(key, 0) * denominator_correction
        try:
            percent_dict[key] = n / d
        except ZeroDivisionError:
            percent_dict[key] = 0

    return percent_dict


class Sh1(Data):

    new_report = 'Свод Приложение-Продажи.xlsx'

    month_list = [['ЯНВАРЬ', 'ФЕВРАЛЬ', 'МАРТ', 'АПРЕЛЬ', 'МАЙ', 'ИЮНЬ',
                   'ИЮЛЬ', 'АВГУСТ', 'СЕНТЯБРЬ', 'ОКТЯБРЬ', 'НОЯБРЬ', 'ДЕКАБРЬ'],
                  [st['white_blue']] * 12]

    head_list = []

    def __init__(self, month, template, file_tags):
        # month -- integer, value of reporting month, must be chosen by user.
        # template -- dict, includes tags and key words to define type of files with raw data.
        # file_tag -- str, key of template.

        super().__init__()
        self.month = month
        self.template = template
        self.file_tags = file_tags
        self.wb = Prepare()

        self.start_column = 2 + (self.month - 1) * len(self.head_list[0])
        self.receiving_data = {}    # Storage of all data used to generate the report
        self.max_row = 0

    def reared(self, key, read_only=True):
        """
        Read file correspond with key of template. Create receiving_data[key].
        :param key: string, key of template correspond to necessary type of file.
        :param read_only: boolean that specified if it is needed to use read-only mode of "OpenPyxls library".
        :return: None
        """
        self.wb.file = self.file_tags[key]
        self.wb.reader(read_only)
        self.receiving_data[key] = {}

    def count_mob(self):
        self.receiving_data['Мобильное приложение'] = {}
        _person_dict = {}
        for city_key in TOWN_GROUP:
            _cache = []
            for value in TOWN_GROUP[city_key]:
                _cache.append(value + ('Мобильное приложение',))
            _person_dict[city_key] = _cache

        person = self.wb.get_data(_person_dict, [1, 4], [7], 'SUM')
        self.receiving_data['Мобильное приложение'] = person
        return self.receiving_data['Мобильное приложение']

    def count_bills(self):
        self.receiving_data['Сумма'] = {}
        bills = self.wb.get_data(TOWN_GROUP, [1, ], [7], 'SUM')
        self.receiving_data['Сумма'] = bills
        return self.receiving_data['Сумма']

    def get_plans(self, target_column):
        """
        Gets data about plane from xlsx-file and enters them to self.receiving_data['Планы'].
        :param target_column: integer, a number of the column, which includes the data about plans.
        :return: None
        """
        self.reared('Планы')
        plans = self.wb.get_data(PLAN_GROUP, [1, ], [target_column, ], 'SUM')

        self.receiving_data['Планы'] = {}
        for k in plans:
            self.receiving_data['Планы'][k] = plans[k]

    def get_period(self):
        first_day_of_month = datetime.today().replace(day=1,
                                                      month=self.month).strftime('%d.%m.%Y')
        today = datetime.today()
        last_day_of_month = (datetime.today().replace(day=1,
                                                      month=(self.month + 1)) - timedelta(days=1))

        if today > last_day_of_month:
            act_day = last_day_of_month.strftime('%d.%m.%Y')
        else:
            act_day = today.strftime('%d.%m.%Y')

        period = [
            ['период',
             '%s' % first_day_of_month,
             '%s' % act_day,
             ],
            [st['white_right'],
             st['white_right'],
             st['white_right'],
             ]
        ]
        return period


class DataSh1(Sh1):
    sheet_name = 'Для рассылки ОД'
    data = [[CITIES_LIST, ],
            [st['smoky20'], ]]
    head_list = [
        ['Заказы через моб.прил. ',
         'Общее количество заказов ',
         'Моб.прил./ Общ. кол-во заказов',
         'План на конец месяца',
         'Выполнение плана по переводу на моб.прил.',
         ],
        [st['apricot20'],
         st['apricot40'],
         st['apricot60'],
         st['apricot70'],
         st['apricot80'],
         ]
    ]

    value = {}
    summary = []
    receiving_data = {}
    chart = False

    def create_template(self):
        self.high = {'1': 27, '2': 40.5, '5': 88.5}
        self.width = {'A': 19.3}
        self.add_width([11, 12, 13, 12, 13.5],
                       start_column=2,
                       count_iter=12)
        self.image = {'A4': {'f_name': os.path.join(IMG_DIR, 'car.png'),
                             'w': 139, 'h': 139}}
        self.rows_group = {}
        self.columns_group = {}
        self.merge = {'A4': 'A5'}
        self.cells = {'A1': ['Свод по количеству реализаций через мобильное приложение'
                             ' от общего количества реализаций по каналу "Доставка"',
                             st['blue20']],
                      'A2': ['*Примечание: Итоги подводятся по всем контрагентам, '
                             'Заказ и реализация ПРОВЕДЕНЫ',
                             st['white_red']]}
        #  Entering the rows with names of cities.
        self.add_cells_by_rows(self.data[0],
                               self.data[1],
                               start_row=6,
                               start_column=1)

        # Entering  the summary row.
        self.max_row = 1 + get_max_row(self.cells.keys())
        self.cells['A%s' % self.max_row] = ['ИТОГО', st['blue20_r']]

    def create_mutable_part(self):
        period = self.get_period()
        #  Entering name of month and merging cells.
        self.add_cells_by_column([self.month_list[0][self.month - 1]],
                                 [self.month_list[1][self.month - 1]],
                                 start_row=3,
                                 start_column=self.start_column)
        self.add_merge('B3', 'F3',
                       count_iter=12,
                       step=1, )

        # Entering data about current reporting period
        self.add_cells_by_column(period[0],
                                 period[1],
                                 start_row=4,
                                 start_column=self.start_column,
                                 count_iter=1, step=1)

        # Entering head of table
        self.add_cells_by_column(self.head_list[0],
                                 self.head_list[1],
                                 start_row=5,
                                 start_column=self.start_column,
                                 count_iter=1,
                                 step=1)

        self.reared('Дисп')
        # Receiving data on the amount of purchases through the mobile application.
        self.count_mob()
        # Receiving data on the total amount of purchases.
        self.count_bills()
        # Counting data on the ratio of the number of purchases through the mobile application to the total.
        self.receiving_data['% от количества'] = {}
        self.receiving_data['% от количества'] = percent(self.receiving_data['Мобильное приложение'],
                                                         self.receiving_data['Сумма'], TOWN_GROUP)

        # Receiving  the planned targets of sales.
        self.get_plans(2)

        # Counting data on the ration of the number of purchases through the mobile application to the target.
        self.receiving_data['% от плана'] = {}
        self.receiving_data['% от плана'] = percent(self.receiving_data['% от количества'],
                                                    self.receiving_data['Планы'],
                                                    TOWN_GROUP)
        # Counting the summary
        sum_mob = sum(self.receiving_data['Мобильное приложение'].values())
        sum_bills = sum(self.receiving_data['Сумма'].values())
        sum_plans = sum(self.receiving_data['Планы'].values())/len(self.receiving_data['Планы'].values())

        DataSh1.summary = [sum_mob,
                           sum_bills,
                           sum_mob / sum_bills,
                           sum_plans,
                           sum_mob/(sum_plans * sum_bills),
                           ]
        summary_style = [st['blue20_r'],
                         st['blue20_r'],
                         st['blue20_red_p'],
                         st['blue20_p'],
                         st['blue20_p'],
                         ]

        DataSh1.value = {key: [[self.receiving_data[key].get(city)
                                for city in CITIES_LIST]
                               ]
                         for key in self.receiving_data}

        # Entering data on the amount of purchases through the mobile application.
        self.add_cells_by_rows(DataSh1.value['Мобильное приложение'],
                               [st['smoky20']],
                               start_row=6,
                               start_column=self.start_column)

        # Entering data on the total amount of purchases.
        self.add_cells_by_rows(DataSh1.value['Сумма'], [st['smoky20']],
                               start_row=6,
                               start_column=self.start_column + 1)

        # Entering data on the ratio of the number
        # of purchases through the mobile application to the total.
        self.add_cells_by_rows(DataSh1.value['% от количества'],
                               [st['smoky20_red_p']],
                               start_row=6, start_column=self.start_column + 2)

        # Entering the planned targets of sales.
        self.add_cells_by_rows(DataSh1.value['Планы'],
                               [st['smoky20_p']],
                               start_row=6, start_column=self.start_column + 3)

        # Entering data on the ration of the number
        # of purchases through the mobile application to the target.
        self.add_cells_by_rows(DataSh1.value['% от плана'],
                               [st['smoky20_p']],
                               start_row=6, start_column=self.start_column + 4)

        # Entering summary.
        if self.max_row == 0:   # If create_template method has not called
            self.max_row = 1 + get_max_row(self.cells.keys())

        self.add_cells_by_column(DataSh1.summary, summary_style,
                                 start_row=self.max_row,
                                 start_column=self.start_column)

        # Merging cells in the first and second rows from the first column
        # to the last filled column.
        last_column = coordinate_from_string(max(self.cells.keys()))[0]
        self.add_merge('A1', '%s1' % last_column)
        self.add_merge('A2', '%s2' % last_column)


class Sh2(Sh1):
    def qr_amount(self):
        qr = self.wb.get_data(CLUSTER_GROUP, [1, 2], [5], 'SUM')
        self.receiving_data['QR'] = qr
        return self.receiving_data['QR']

    def count_bills(self):
        last_column = self.wb.sheet_r.max_column
        column_list = list(range(4, last_column))

        amount = self.wb.get_data(CLUSTER_GROUP, [1, 2], column_list, 'COUNT')
        self.receiving_data['Продажи'] = amount

        return self.receiving_data['Продажи']


class DataSh2(Sh2):
    sheet_name = 'Для рассылки ФМ'
    data = [[CITIES_LIST], [st['smoky20']]]
    head_list = [['Счит. QR', 'Чеки', 'QR / Чеки', 'План на конец месяца', '% выполнения плана'],
                 [st['apricot20'], st['apricot40'], st['apricot60'], st['apricot70'], st['apricot80'], ]]

    value = {}
    summary = []
    receiving_data = {}
    chart = False

    def create_template(self):
        self.high = {'1': 57, '2': 42, '5': 81}
        self.width = {'A': 18, }
        self.image = {'A4': {'f_name': os.path.join(IMG_DIR, 'car.png'), 'w': 129, 'h': 129}}
        self.rows_group = dict([(12, 14), ])
        self.columns_group = {}

        #  Entering name of month and merging cells.
        self.cells = {'A1': ['Свод по количеству считываний QR кода и количеству чеков в ФМ по территориям ',
                             st['blue20']],
                      'A2': ['* Примечание: если клиент в день совершил несколько покупок '
                             'он засчитывается один раз ',
                             st['white_red']]}
        self.merge = {'A3': 'A4'}

        #  Entering the rows with names of cities.
        self.add_cells_by_rows(self.data[0], self.data[1], start_row=6, start_column=1)

        self.max_row = 1 + get_max_row(self.cells.keys())

        #  Entering the rows with names of sub-clusters (another style).
        self.add_cells_by_rows([SUB_CLUSTER_LIST], [st['white_right']],
                               start_row=self.max_row, start_column=1)

        # Entering  the summary row.
        self.max_row = 1 + get_max_row(self.cells.keys())
        self.cells['A%s' % self.max_row] = ['ИТОГО', st['blue20_r']]

    def create_mutable_part(self):
        period = self.get_period()

        #  Entering name of month and merging cells.
        self.add_cells_by_column([self.month_list[0][self.month - 1]],
                                 [self.month_list[1][self.month - 1]],
                                 start_row=3, start_column=self.start_column)
        self.add_merge('B3', 'F3', count_iter=12, step=1, )

        # Entering data about current reporting period
        self.add_cells_by_column(period[0],
                                 period[1],
                                 start_row=4,
                                 start_column=self.start_column,
                                 count_iter=1, step=1)

        # Entering head of table
        self.add_cells_by_column(self.head_list[0], self.head_list[1], start_row=5,
                                 start_column=self.start_column,
                                 count_iter=1, step=1)

        # Receiving data on the amount of QR-codes activations.
        self.reared('QR')
        self.qr_amount()

        # Receiving data on the total amount of purchases.
        self.reared('Продажи', read_only=False)
        self.count_bills()

        # Receiving the planned targets of sales.
        self.get_plans(3)

        # Counting the summary
        sum_qr = sum(self.receiving_data['QR'].values())
        sum_bills = sum(self.receiving_data['Продажи'].values())
        sum_plans = sum(self.receiving_data['Планы'].values()) / len(self.receiving_data['Планы'].values())

        DataSh2.summary = [sum_qr,
                           sum_bills,
                           sum_qr / sum_bills,
                           sum_plans,
                           sum_qr / (sum_bills * sum_plans),
                           ]
        summary_style = [st['blue20_r'],
                         st['blue20_r'],
                         st['blue20_p'],
                         st['blue20_p'],
                         st['blue20_p'],
                         ]

        # Counting the total values from sub-cluster.
        for tag in self.receiving_data.keys():
            for city in FULL_CITIES_LIST:
                _cache = 0
                for clust in self.receiving_data[tag]:
                    if city in clust:
                        _cache += self.receiving_data[tag][clust]
                self.receiving_data[tag][city] = _cache

        # Counting data on the ratio of the number of QR-codes activations
        # to the total amount of purchases.
        self.receiving_data['QR/Чеки'] = percent(self.receiving_data.get('QR'),
                                                 self.receiving_data.get('Продажи'),
                                                 self.receiving_data.get('QR'))

        # Counting data on the ration of the number
        # of purchases through the mobile application to the target.
        self.receiving_data['% от плана'] = {}
        self.receiving_data['% от плана'] = percent(self.receiving_data['QR/Чеки'],
                                                    self.receiving_data['Планы'],
                                                    PLAN_GROUP)

        DataSh2.value = {key: [self.receiving_data.get(key).get(city, 0)
                               for city in FULL_CITIES_LIST] for key in self.receiving_data}

        # Entering data on the amount of QR-codes activations (without sub-clusters).
        self.add_cells_by_rows([DataSh2.value['QR'][:-3]],
                               self.data[1],
                               start_row=6,
                               start_column=self.start_column)

        # Entering data on the amount of QR-codes activations (only sub-clusters).
        self.add_cells_by_rows([DataSh2.value['QR'][-3:]],
                               [st['white_right']],
                               start_row=12,
                               start_column=self.start_column)

        # Entering data on the total amount of purchases (without sub-clusters).
        self.add_cells_by_rows([DataSh2.value['Продажи'][:-3]],
                               self.data[1],
                               start_row=6,
                               start_column=self.start_column+1)

        # Entering data on the total amount of purchases (only sub-clusters).
        self.add_cells_by_rows([DataSh2.value['Продажи'][-3:]],
                               [st['white_right']],
                               start_row=12,
                               start_column=self.start_column+1)

        # Entering data on the ratio of the number of QR-codes activations
        # to the total amount of purchases (without sub-clusters).
        self.add_cells_by_rows([DataSh2.value['QR/Чеки'][:-3]],
                               [st['smoky20_p']],
                               start_row=6,
                               start_column=self.start_column+2)

        # Entering data on the ratio of the number of QR-codes activations
        # to the total amount of purchases (only sub-clusters).
        self.add_cells_by_rows([DataSh2.value['QR/Чеки'][-3:]],
                               [st['white_right_p']],
                               start_row=12,
                               start_column=self.start_column+2)

        # Entering the planned targets of sales (without sub-clusters).
        self.add_cells_by_rows([DataSh2.value['Планы'][:-3]],
                               [st['smoky20_p']],
                               start_row=6,
                               start_column=self.start_column+3)

        # Entering the planned targets of sales (only sub-clusters).
        self.add_cells_by_rows([DataSh2.value['Планы'][-3:]],
                               [st['white_right_p']],
                               start_row=12,
                               start_column=self.start_column+3)

        # Entering data on the ration of the number
        # of purchases through the mobile application to the target (without sub-clusters).
        self.add_cells_by_rows([DataSh2.value['% от плана'][:-3]],
                               [st['smoky20_p']],
                               start_row=6,
                               start_column=self.start_column+4)
        # Entering data on the ration of the number
        # of purchases through the mobile application to the target (only sub-clusters).
        self.add_cells_by_rows([DataSh2.value['% от плана'][-3:]],
                               [st['white_right_p']],
                               start_row=12,
                               start_column=self.start_column+4)
        # Entering the summary
        if self.max_row == 0:   # If create_template method has not called
            self.max_row = 1 + get_max_row(self.cells.keys())

        self.add_cells_by_column(DataSh2.summary,
                                 summary_style,
                                 start_row=self.max_row,
                                 start_column=self.start_column)

        # Merging cells in the first and second rows from the first column
        # to the last filled column.
        last_column = coordinate_from_string(max(self.cells.keys()))[0]
        self.add_merge('A1', '%s1' % last_column)
        self.add_merge('A2', '%s2' % last_column)


class DataSh3(DataSh1):
    sheet_name = 'Итоговый ОД'

    data = [[CITIES_LIST],
            [st['smoky20_v']]]
    head_list = [['Город', 'Период', 'Показатель'] + Sh1.month_list[0],
                 [st['purple60']]*15]
    chart = True
    chart_title = 'Приложение/Все заказы'

    def __init__(self, month, template, file_tags, year, last_report, last_year_report):
        super().__init__(month, template, file_tags)
        self.year = year
        self.last_year_report = last_year_report
        self.start_column = 4 + (self.month - 1)
        self.last_report = last_report

    def enter_indicators(self, data, stiles, start_row, start_column, step, length):
        for ind, stile in zip(data, stiles):
            self.add_cells_by_rows([[ind]*length],
                                   [stile],
                                   start_row=start_row,
                                   start_column=start_column,
                                   step=step)
            start_row += 1

    def create_template(self):
        self.high = {'1': 24, '2': 176}
        self.width = {'A': 24, 'B': 12, 'C': 16}
        self.rows_group = dict([(4, 5), (10, 11), (13, 14), (19, 20), (22, 23), (28, 29), (31, 32), (37, 38), (40, 41),
                                (46, 47), (49, 50), (55, 56), (58, 59), (64, 65)])
        self.merge = {'A1': 'O1'}

        self.cells['A1'] = ['СРАВНИТЕЛЬНЫЙ СВОД ПО СЧИТЫВАНИЯМ (Доставка)',
                            st['blue40_white']]

        # Entering the head of the table.
        self.add_cells_by_column(self.head_list[0],
                                 self.head_list[1],
                                 start_row=3,
                                 start_column=1)

        #  Entering the rows with the names of cities.
        self.add_cells_by_rows(self.data[0],
                               self.data[1],
                               start_row=4,
                               start_column=1,
                               step=9)

        # Entering the summary row.
        self.cells['A58'] = ['ИТОГО',
                             st['blue20']]

        # Merging the cells with the city names and the summary.
        self.add_merge(start_merge='A4',
                       end_merge='A11',
                       step=2,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicator 'last year'.
        last_year = [[['%s' % (self.year-1)] * 7],
                     [st['blue40_white_11']]]
        current_year = [[['%s' % self.year] * 7],
                        [st['apricot60']]]
        compare_years = [[['%s/%s' % (self.year, (self.year-1))] * 7],
                         [st['blue05']]]

        self.add_cells_by_rows(last_year[0],
                               last_year[1],
                               start_row=4,
                               start_column=2,
                               step=9)

        self.add_merge(start_merge='B4',
                       end_merge='B6',
                       step=7,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicator 'current year'.
        self.add_cells_by_rows(current_year[0],
                               current_year[1],
                               start_row=7,
                               start_column=2,
                               step=9)

        self.add_merge(start_merge='B7',
                       end_merge='B9',
                       step=7,
                       count_iter=7,
                       direction='downward')
        # Entering data on the indicator 'last year / current year'.
        self.add_cells_by_rows(compare_years[0],
                               compare_years[1],
                               start_row=10,
                               start_column=2,
                               step=9)

        self.add_merge(start_merge='B10',
                       end_merge='B11',
                       step=8,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicators.
        indicator_list = [
            'Приложение',
            'Заказы',
            'Прил. / заказы',
            'Приложение',
            'Заказы',
            'Прил. / заказы',
            'Приложение',
            'Заказы',
        ]

        indicator_styles = [
            st['blue05_l'],
            st['blue10_l'],
            st['blue40_white_l'],
            st['apricot20_l'],
            st['apricot40_l'],
            st['apricot60_l'],
            st['white_blue_l'],
            st['white_blue_l'],
        ]

        self.enter_indicators(indicator_list,
                              indicator_styles,
                              start_row=4,
                              start_column=3,
                              step=9,
                              length=7)

        # Entering data from last year report.
        last_year_obj = Prepare()
        last_year_obj.reader(sheet='Итоговый ОД',
                             file_name=self.last_year_report)
        for i in range(7):
            last_data = last_year_obj.get_data_by_number(start_row=(7+(i*9)),
                                                         max_row=(9+(i*9)),
                                                         start_column=4,
                                                         max_column=15)
            last_data_st = [[st['blue05_r']]*12,
                            [st['blue10_r']]*12,
                            [st['blue40_white_11_p']]*12]

            self.add_cells_by_column(last_data[0],
                                     last_data_st[0],
                                     start_row=(4+(i*9)),
                                     start_column=4)
            self.add_cells_by_column(last_data[1],
                                     last_data_st[1],
                                     start_row=(5+(i*9)),
                                     start_column=4)
            self.add_cells_by_column(last_data[2],
                                     last_data_st[2],
                                     start_row=(6+(i*9)),
                                     start_column=4)

    def create_mutable_part(self):

        data_st = [[st['apricot20_r']],
                   [st['apricot40_r']],
                   [st['apricot60_p']]]

        # Entering data from DataSh1.
        mob = self.value['Мобильное приложение'][0]
        summ = self.value['Сумма'][0]
        ratio = self.value['% от количества'][0]

        mob.append(self.summary[0])
        summ.append(self.summary[1])
        ratio.append(self.summary[2])

        # Entering data on the amount of purchases through the mobile application.
        self.add_cells_by_rows([mob],
                               data_st[0],
                               start_row=7,
                               start_column=self.start_column,
                               step=9)
        # Entering data on the total amount of purchases.
        self.add_cells_by_rows([summ],
                               data_st[1],
                               start_row=8,
                               start_column=self.start_column,
                               step=9)

        # Entering data on the ratio of the number
        # of purchases through the mobile application to the total.
        self.add_cells_by_rows([ratio],
                               data_st[2],
                               start_row=9,
                               start_column=self.start_column,
                               step=9)

        # Getting and entering data from last report
        # to compare with current year indicators.
        last_data = Prepare()
        last_data.reader(sheet=self.sheet_name,
                         read_only=False,
                         file_name=self.last_report)
        for i in range(len(mob)):
            last_app = last_data.get_data_by_number(start_row=4+9*i,
                                                    max_row=4+9*i,
                                                    start_column=self.start_column,
                                                    max_column=self.start_column)
            last_orders = last_data.get_data_by_number(start_row=5+9*i,
                                                       max_row=5+9*i,
                                                       start_column=self.start_column,
                                                       max_column=self.start_column)

            compare_app = (mob[i])/last_app
            compare_orders = (summ[i])/last_orders

            if compare_app > 1:
                app_st = [st['good_p']]
            else:
                app_st = [st['bad_p']]

            if compare_orders > 1:
                ord_st = [st['good_p']]
            else:
                ord_st = [st['bad_p']]

            self.add_cells_by_rows([[compare_app]],
                                   app_st,
                                   start_row=10+9*i,
                                   start_column=self.start_column)
            self.add_cells_by_rows([[compare_orders]],
                                   ord_st,
                                   start_row=11 + 9 * i,
                                   start_column=self.start_column)


class DataSh4(DataSh3):

    sheet_name = 'Итоговый ФМ'
    chart = True
    chart_title = 'QR/Чеки'

    def create_template(self):
        self.high = {'1': 24, '2': 176}
        self.width = {'A': 24, 'B': 12, 'C': 16}
        self.rows_group = dict(
            [(4, 5), (10, 11), (13, 14), (19, 20), (22, 23), (28, 29), (31, 32), (37, 38), (40, 41),
             (46, 47), (49, 50), (55, 56), (58, 59), (64, 65)])
        self.merge = {'A1': 'O1'}

        self.cells['A1'] = ['СРАВНИТЕЛЬНЫЙ СВОД ПО СЧИТЫВАНИЯМ (Обменные пункты)',
                            st['blue40_white']]

        # Entering the head of the table.
        self.add_cells_by_column(self.head_list[0],
                                 self.head_list[1],
                                 start_row=3,
                                 start_column=1)

        #  Entering the rows with the names of cities.
        self.add_cells_by_rows(self.data[0],
                               self.data[1],
                               start_row=4,
                               start_column=1,
                               step=9)

        # Entering the summary row.
        self.cells['A58'] = ['ИТОГО',
                             st['blue20']]

        # Merging the cells with the city names and the summary.
        self.add_merge(start_merge='A4',
                       end_merge='A11',
                       step=2,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicator 'last year'.
        last_year = [[['%s' % (self.year - 1)] * 7],
                     [st['blue40_white_11']]]
        current_year = [[['%s' % self.year] * 7],
                        [st['apricot60']]]
        compare_years = [[['%s/%s' % (self.year, (self.year - 1))] * 7],
                         [st['blue05']]]

        self.add_cells_by_rows(last_year[0],
                               last_year[1],
                               start_row=4,
                               start_column=2,
                               step=9)

        self.add_merge(start_merge='B4',
                       end_merge='B6',
                       step=7,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicator 'current year'.
        self.add_cells_by_rows(current_year[0],
                               current_year[1],
                               start_row=7,
                               start_column=2,
                               step=9)
        self.add_merge(start_merge='B7',
                       end_merge='B9',
                       step=7,
                       count_iter=7,
                       direction='downward')

        # Entering data on the indicator 'last year / current year'.
        self.add_cells_by_rows(compare_years[0],
                               compare_years[1],
                               start_row=10,
                               start_column=2,
                               step=9)

        self.add_merge(start_merge='B10',
                       end_merge='B11',
                       step=8,
                       count_iter=7,
                       direction='downward')
        # Entering data on the indicators.
        indicator_list = [
            'QR',
            'Чеки',
            'QR/чеки',
            'QR',
            'Чеки',
            'QR/чеки',
            'QR',
            'Чеки',
        ]
        indicator_styles = [
            st['blue05_l'],
            st['blue10_l'],
            st['blue40_white_l'],
            st['apricot20_l'],
            st['apricot40_l'],
            st['apricot60_l'],
            st['white_blue_l'],
            st['white_blue_l']
        ]
        self.enter_indicators(indicator_list,
                              indicator_styles,
                              start_row=4,
                              start_column=3,
                              step=9,
                              length=7)

        # Entering data from last year report.
        last_year_obj = Prepare()
        last_year_obj.reader(sheet='Итоговый ФМ',
                             file_name=self.last_year_report)

        for i in range(7):
            last_data = last_year_obj.get_data_by_number(start_row=(7+(i*9)),
                                                         max_row=(9+(i*9)),
                                                         start_column=4,
                                                         max_column=15)

            last_data_st = [[st['blue05_r']] * 12,
                            [st['blue10_r']] * 12,
                            [st['blue40_white_11_p']] * 12]

            self.add_cells_by_column(last_data[0],
                                     last_data_st[0],
                                     start_row=(4 + (i * 9)),
                                     start_column=4)
            self.add_cells_by_column(last_data[1],
                                     last_data_st[1],
                                     start_row=(5 + (i * 9)),
                                     start_column=4)
            self.add_cells_by_column(last_data[2],
                                     last_data_st[2],
                                     start_row=(6 + (i * 9)),
                                     start_column=4)

    def create_mutable_part(self):

        data_st = [[st['apricot20_r']],
                   [st['apricot40_r']],
                   [st['apricot60_p']]]

        # Receiving data from DataSh2.
        self.value = DataSh2.value
        self.summary = DataSh2.summary

        self.value['QR'].append(self.summary[0])
        self.value['Продажи'].append(self.summary[1])
        self.value['QR/Чеки'].append(self.summary[2])

        # Entering data on the amount of QR-codes activations by each city.
        self.add_cells_by_rows([self.value['QR'][:-4]],
                               data_st[0],
                               start_row=7,
                               start_column=self.start_column,
                               step=9)

        # Entering data on the total amount of QR-codes activations.
        self.add_cells_by_rows([self.value['QR'][-1:]],
                               data_st[0],
                               start_row=61,
                               start_column=self.start_column,
                               step=1)

        # Entering data on the amount of purchases by each city.
        self.add_cells_by_rows([self.value['Продажи'][:-4]],
                               data_st[1],
                               start_row=8,
                               start_column=self.start_column,
                               step=9)

        # Entering data on the total amount of purchases.
        self.add_cells_by_rows([self.value['Продажи'][-1:]],
                               data_st[1],
                               start_row=62,
                               start_column=self.start_column,
                               step=1)
        # Entering data on the ratio of the number of QR-codes activations
        # to the purchases by each city
        self.add_cells_by_rows([self.value['QR/Чеки'][:-4]],
                               data_st[2],
                               start_row=9,
                               start_column=self.start_column,
                               step=9)
        # Entering data on the ratio of the total number of QR-codes activations
        # to the purchases.
        self.add_cells_by_rows([self.value['QR/Чеки'][-1:]],
                               data_st[2],
                               start_row=63,
                               start_column=self.start_column,
                               step=9)

        # Getting and entering data from last report
        # to compare with current year indicators.
        last_data = Prepare()
        last_data.file = self.last_report
        last_data.reader(sheet=self.sheet_name,
                         read_only=False)

        for i in range(len(self.value['QR'][:-3])):

            last_qr = last_data.get_data_by_number(start_row=4 + 9 * i,
                                                   max_row=4 + 9 * i,
                                                   start_column=self.start_column,
                                                   max_column=self.start_column)
            last_orders = last_data.get_data_by_number(start_row=5 + 9 * i,
                                                       max_row=5 + 9 * i,
                                                       start_column=self.start_column,
                                                       max_column=self.start_column)

            compare_qr = (self.value['QR'][i]) / last_qr
            compare_orders = (self.value['Продажи'][i]) / last_orders

            if compare_qr > 1:
                app_st = [st['good_p']]
            else:
                app_st = [st['bad_p']]

            if compare_orders > 1:
                ord_st = [st['good_p']]
            else:
                ord_st = [st['bad_p']]

            self.add_cells_by_rows([[compare_qr]], app_st,
                                   start_row=10 + 9 * i,
                                   start_column=self.start_column)
            self.add_cells_by_rows([[compare_orders]],
                                   ord_st, start_row=11 + 9 * i,
                                   start_column=self.start_column)
